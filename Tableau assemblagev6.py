import streamlit as st
import pandas as pd
import numpy as np
from datetime import datetime
import tempfile
import hashlib
import json
import time
import secrets
import re
from urllib.parse import urlencode

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ====== Supabase + live dashboard deps ======
from supabase import create_client
import plotly.graph_objects as go
from streamlit_autorefresh import st_autorefresh

# ==========================================================
# CONFIG / UI
# ==========================================================
st.set_page_config(page_title="🧪 Tableau Assemblage + Stocks", page_icon="🧪", layout="wide")

st.markdown(
    """
    <style>
      .block-container {padding-top: 1.2rem;}
      div[data-testid="stExpander"] details summary {font-size: 1.03rem;}
      .pill {display:inline-block; padding:.2rem .55rem; border-radius:999px; background:#f1f3f5; margin-right:.35rem;}
      .small {opacity:.75; font-size:.9rem;}
      .card {padding: .8rem 1rem; border-radius: 14px; background:#f8f9fa; border: 1px solid #e9ecef;}
      code {font-size: .9rem;}
      .mono {font-family: ui-monospace, SFMono-Regular, Menlo, Monaco, Consolas, "Liberation Mono", "Courier New", monospace;}
    </style>
    """,
    unsafe_allow_html=True
)

st.title("🧪 Tableau Assemblage — avec gestion de stock + dégustation live")
st.markdown(
    """
    <span class="pill">C/N/M conservés</span>
    <span class="pill">Essais multiples</span>
    <span class="pill">Récap % fiable</span>
    <span class="pill">Stocks décrémentés par Code Produit</span>
    <span class="pill">Anti double-application</span>
    <span class="pill">Dégustation live (Supabase + PIN)</span>
    <span class="pill">Comparaison cuves</span>
    """,
    unsafe_allow_html=True
)

CEPAGE_LABEL = {"C": "Chardonnay", "N": "Pinot Noir", "M": "Meunier"}

# ==========================================================
# SUPABASE (Degustation) - helpers
# ==========================================================
@st.cache_resource
def sb():
    # Requiert dans Streamlit Secrets :
    # SUPABASE_URL="..."
    # SUPABASE_SERVICE_ROLE_KEY="..."
    url = st.secrets.get("SUPABASE_URL", "")
    key = st.secrets.get("SUPABASE_SERVICE_ROLE_KEY", "")
    if not url or not key:
        return None
    return create_client(url, key)

def supabase_ready():
    return sb() is not None

def sha256_hex(s: str) -> str:
    return hashlib.sha256(s.encode("utf-8")).hexdigest()

def make_pin(n_digits: int = 4) -> str:
    # 4-digit PIN, sans 0000
    while True:
        pin = "".join(str(secrets.randbelow(10)) for _ in range(n_digits))
        if pin != "0" * n_digits:
            return pin

def normalize_pin(pin: str) -> str:
    if pin is None:
        return ""
    return re.sub(r"\D", "", str(pin))[:6]

def sup_create_essai(nom: str, cuves: list[str]) -> tuple[str, str]:
    """
    Crée un essai et génère un PIN.
    Retourne (essai_id, pin_en_clair).
    """
    pin = make_pin(4)
    pin_hash = sha256_hex(pin)

    res = sb().table("essais").insert({"nom": nom, "cuves": cuves, "pin_hash": pin_hash}).execute()
    essai_id = res.data[0]["id"]
    return essai_id, pin

def sup_list_essais(limit: int = 50) -> pd.DataFrame:
    res = sb().table("essais").select("id,created_at,nom,cuves,pin_hash").order("created_at", desc=True).limit(limit).execute()
    return pd.DataFrame(res.data)

def sup_get_essai(essai_id: str) -> dict:
    res = sb().table("essais").select("id,created_at,nom,cuves,pin_hash").eq("id", essai_id).single().execute()
    return res.data

def sup_check_pin(essai: dict, pin: str) -> bool:
    pin = normalize_pin(pin)
    if not pin:
        return False
    ph = (essai or {}).get("pin_hash") or ""
    if not ph:
        # si jamais pas de pin_hash, on autorise (mais ce n'est pas recommandé)
        return True
    return sha256_hex(pin) == ph

def sup_upsert_note_with_retry(
    essai_id: str,
    cuve: str,
    degustateur: str,
    notes: dict,
    commentaire: str,
    tries: int = 3
):
    """
    Upsert via on_conflict (plus fiable que insert/except update).
    + retry léger pour les instabilités réseau.
    """
    last_err = None
    row = {
        "essai_id": essai_id,
        "cuve": cuve,
        "degustateur": degustateur,
        **notes,
        "commentaire": commentaire or "",
    }

    for k in range(tries):
        try:
            sb().table("notes").upsert(row, on_conflict="essai_id,cuve,degustateur").execute()
            return True, None
        except Exception as e:
            last_err = e
            time.sleep(0.35 * (k + 1))  # backoff simple
    return False, last_err

@st.cache_data(ttl=2, show_spinner=False)
def sup_fetch_notes(essai_id: str) -> pd.DataFrame:
    res = (
        sb().table("notes")
        .select("created_at,essai_id,cuve,degustateur,acidite,amertume,mineralite,volume,sucrosite,defaut,commentaire")
        .eq("essai_id", essai_id)
        .order("created_at", desc=False)
        .limit(5000)
        .execute()
    )
    df = pd.DataFrame(res.data)
    if df.empty:
        return df
    for c in ["acidite", "amertume", "mineralite", "volume", "sucrosite", "defaut"]:
        df[c] = pd.to_numeric(df[c], errors="coerce")
    return df

RADAR_AXES = ["Acidité", "Amertume", "Minéralité", "Volume", "Sucrosité", "Pureté"]  # Pureté = 6 - défaut

def radar_fig(df_cuve: pd.DataFrame, by_taster: bool = False):
    d = df_cuve.copy()
    d["purete"] = 6 - d["defaut"]

    mean_vals = [
        d["acidite"].mean(),
        d["amertume"].mean(),
        d["mineralite"].mean(),
        d["volume"].mean(),
        d["sucrosite"].mean(),
        d["purete"].mean(),
    ]

    fig = go.Figure()
    fig.add_trace(
        go.Scatterpolar(
            r=mean_vals + [mean_vals[0]],
            theta=RADAR_AXES + [RADAR_AXES[0]],
            fill="toself",
            name="Moyenne",
        )
    )

    if by_taster:
        for degust, g in d.groupby("degustateur"):
            vals = [
                g["acidite"].mean(),
                g["amertume"].mean(),
                g["mineralite"].mean(),
                g["volume"].mean(),
                g["sucrosite"].mean(),
                (6 - g["defaut"]).mean(),
            ]
            fig.add_trace(
                go.Scatterpolar(
                    r=vals + [vals[0]],
                    theta=RADAR_AXES + [RADAR_AXES[0]],
                    name=str(degust),
                )
            )

    fig.update_layout(
        margin=dict(l=10, r=10, t=25, b=10),
        polar=dict(radialaxis=dict(visible=True, range=[1, 5], dtick=1)),
        height=300,
        showlegend=by_taster,
    )
    return fig

def build_scores_table(df: pd.DataFrame) -> pd.DataFrame:
    """
    Table moyenne par cuve (agrège tous dégustateurs).
    Colonnes: acidite, amertume, mineralite, volume, sucrosite, purete
    """
    if df.empty:
        return pd.DataFrame()

    d = df.copy()
    d["purete"] = 6 - d["defaut"]

    scores = (
        d.groupby("cuve", as_index=False)
        .agg(
            acidite=("acidite", "mean"),
            amertume=("amertume", "mean"),
            mineralite=("mineralite", "mean"),
            volume=("volume", "mean"),
            sucrosite=("sucrosite", "mean"),
            purete=("purete", "mean"),
            votes=("cuve", "size"),
            degustateurs=("degustateur", "nunique"),
        )
    )

    # arrondis lisibles
    for c in ["acidite", "amertume", "mineralite", "volume", "sucrosite", "purete"]:
        scores[c] = scores[c].round(2)

    # score global simple (moyenne des 6)
    scores["score_global"] = scores[["acidite", "amertume", "mineralite", "volume", "sucrosite", "purete"]].mean(axis=1).round(2)
    scores = scores.sort_values(["score_global", "votes"], ascending=[False, False])
    return scores

def plot_heatmap_scores(scores: pd.DataFrame):
    # Heatmap (cuves x critères)
    if scores.empty:
        return None

    criteria = ["acidite", "amertume", "mineralite", "volume", "sucrosite", "purete"]
    z = scores[criteria].to_numpy()
    y = scores["cuve"].astype(str).tolist()
    x = ["Acidité", "Amertume", "Minéralité", "Volume", "Sucrosité", "Pureté"]

    fig = go.Figure(
        data=go.Heatmap(
            z=z,
            x=x,
            y=y,
            zmin=1,
            zmax=5,
            colorbar=dict(title="Note"),
        )
    )
    fig.update_layout(height=min(900, 120 + 28 * len(y)), margin=dict(l=10, r=10, t=35, b=10))
    return fig

def plot_distance_to_reference(scores: pd.DataFrame, ref_cuve: str):
    """
    Distance euclidienne sur 6 critères (plus petit = plus proche de la cuve de référence).
    """
    if scores.empty or not ref_cuve or ref_cuve not in set(scores["cuve"]):
        return None

    criteria = ["acidite", "amertume", "mineralite", "volume", "sucrosite", "purete"]
    m = scores.set_index("cuve")[criteria]
    ref = m.loc[ref_cuve].to_numpy()
    dist = ((m.to_numpy() - ref) ** 2).sum(axis=1) ** 0.5
    dist = pd.Series(dist, index=m.index).sort_values()

    fig = go.Figure(
        data=go.Bar(
            x=dist.values,
            y=dist.index.astype(str).tolist(),
            orientation="h",
        )
    )
    fig.update_layout(
        height=min(900, 120 + 26 * len(dist)),
        margin=dict(l=10, r=10, t=35, b=10),
        xaxis_title="Distance (écart global)",
        yaxis_title="Cuve",
        title=f"Écart global vs {ref_cuve} (6 critères)",
    )
    return fig

# ==========================================================
# HELPERS COMMUNS
# ==========================================================
def norm_str_series(s: pd.Series) -> pd.Series:
    return s.astype(str).str.strip()

def normalize_str(x):
    if pd.isna(x):
        return ""
    return str(x).strip()

def normalize_cuve_number(x):
    """Corrige des cuves typées en décimal (0,0351) -> 351."""
    if pd.isna(x):
        return x
    if isinstance(x, str):
        xs = x.strip().replace(" ", "").replace(",", ".")
        try:
            x = float(xs)
        except Exception:
            return x
    try:
        v = float(x)
    except Exception:
        return x
    if 0 < v < 1:
        v = v * 10000
    return int(round(v))

def cuve_to_int_or_none(x):
    """Convertit N° Cuve en int si possible, sinon None."""
    if pd.isna(x):
        return None
    try:
        if isinstance(x, str):
            xs = x.strip().replace(" ", "").replace(",", ".")
            if xs == "":
                return None
            x = xs
        v = pd.to_numeric(x, errors="coerce")
        if pd.isna(v):
            return None
        return int(round(float(v)))
    except Exception:
        return None

def to_cepage_code(v):
    if pd.isna(v):
        return ""
    s = str(v).strip()
    su = s.upper()
    if su in ("C", "N", "M"):
        return su
    if su.startswith("CHARD"):
        return "C"
    if "PINOT" in su and "NOIR" in su:
        return "N"
    if "MEUNIER" in su:
        return "M"
    return su

def essai_cols(e: int):
    return {
        "vol": f"Volume (L) E{e}",
        "solde": f"Solde E{e}",
        "qty": f"Quantité utilisée E{e}",
        "pct": f"% E{e}",
        "c250": f"250 E{e}",
        "c500": f"500 E{e}",
    }

def excel_cols_for_essai(e: int, start_base_col: int = 7):
    start = start_base_col + (e - 1) * 6
    return {"start": start, "vol": start, "solde": start + 1, "qty": start + 2, "pct": start + 3, "c250": start + 4, "c500": start + 5, "end": start + 5}

def force_excel_years_to_int(ws, year_col_letter: str, cuve_col_letter: str, start_row: int, end_row: int):
    for r in range(start_row, end_row + 1):
        cuv = ws[f"{cuve_col_letter}{r}"].value
        if cuv in (None, "", 0):
            continue
        v = ws[f"{year_col_letter}{r}"].value
        if isinstance(v, str):
            vv = v.strip()
            if vv.isdigit():
                ws[f"{year_col_letter}{r}"].value = int(vv)

def coerce_float(series):
    return pd.to_numeric(series, errors="coerce")

def find_header_row(excel_path, needle="Clé Produit en Cuve", sheet_name=0, max_scan=80):
    raw = pd.read_excel(excel_path, header=None, sheet_name=sheet_name)
    max_r = min(max_scan, len(raw))
    for i in range(max_r):
        row = raw.iloc[i].astype(str)
        if row.str.contains(needle, na=False).any():
            return i
    return None

# ==========================================================
# STOCK: anti double-application + delta
# ==========================================================
def make_fingerprint(file_bytes: bytes, essai: str, ref: str, date_conso) -> str:
    h = hashlib.sha256()
    h.update(file_bytes)
    h.update(str(essai).encode("utf-8"))
    h.update(str(ref).encode("utf-8"))
    h.update(str(date_conso).encode("utf-8"))
    return h.hexdigest()[:16]

def journal_has_fingerprint(journal_df: pd.DataFrame, fingerprint: str) -> bool:
    if journal_df is None or journal_df.empty:
        return False
    if "Fingerprint" not in journal_df.columns:
        return False
    return (journal_df["Fingerprint"].astype(str) == str(fingerprint)).any()

def build_delta_table(snapshot_df: pd.DataFrame, ledger_df: pd.DataFrame) -> pd.DataFrame:
    s = snapshot_df.copy()
    l = ledger_df.copy()

    s["Code Produit en Cuve"] = s["Code Produit en Cuve"].astype(str).str.strip()
    l["Code Produit en Cuve"] = l["Code Produit en Cuve"].astype(str).str.strip()

    if "Stock_Etat_L" not in s.columns:
        raise ValueError("snapshot_df doit contenir Stock_Etat_L")
    if "Stock restant (L)" not in l.columns:
        raise ValueError("ledger_df doit contenir Stock restant (L)")

    delta = s.merge(
        l[["Code Produit en Cuve", "Stock restant (L)"]],
        on="Code Produit en Cuve",
        how="outer"
    )

    delta["Stock_Etat_L"] = pd.to_numeric(delta["Stock_Etat_L"], errors="coerce").fillna(0.0)
    delta["Stock restant (L)"] = pd.to_numeric(delta["Stock restant (L)"], errors="coerce").fillna(0.0)
    delta["Écart (Etat - Ledger)"] = (delta["Stock_Etat_L"] - delta["Stock restant (L)"]).round(2)

    delta["__abs"] = delta["Écart (Etat - Ledger)"].abs()
    delta = delta.sort_values("__abs", ascending=False).drop(columns="__abs")
    return delta

# ==========================================================
# ONGLET 3 : STOCK UPDATE
# ==========================================================
def build_stock_snapshot(df_stock):
    required = {"Produit", "En Stock"}
    if not required.issubset(set(df_stock.columns)):
        missing = sorted(list(required - set(df_stock.columns)))
        raise ValueError(f"Colonnes manquantes dans l'état de stock: {missing}")

    df = df_stock.copy()
    df["Produit"] = df["Produit"].apply(normalize_str)
    df["En Stock"] = coerce_float(df["En Stock"]).fillna(0)

    agg = (
        df.groupby("Produit", as_index=False)
        .agg(Stock_Etat_L=("En Stock", "sum"))
        .sort_values("Produit")
    )
    agg.rename(columns={"Produit": "Code Produit en Cuve"}, inplace=True)
    return agg

def init_ledger_from_snapshot(snapshot_df):
    led = snapshot_df.copy()
    led["Stock initial (L)"] = led["Stock_Etat_L"]
    led["Consommé cumul (L)"] = 0.0
    led["Stock restant (L)"] = led["Stock initial (L)"] - led["Consommé cumul (L)"]
    led = led[["Code Produit en Cuve", "Stock initial (L)", "Consommé cumul (L)", "Stock restant (L)"]]
    return led

def read_ledger(ledger_xlsx):
    df = pd.read_excel(ledger_xlsx, sheet_name="STOCK_MAJ" if "STOCK_MAJ" in pd.ExcelFile(ledger_xlsx).sheet_names else 0)
    required = {"Code Produit en Cuve", "Stock initial (L)", "Consommé cumul (L)", "Stock restant (L)"}
    if not required.issubset(set(df.columns)):
        raise ValueError(
            "Ledger invalide. Colonnes attendues : Code Produit en Cuve, Stock initial (L), Consommé cumul (L), Stock restant (L)"
        )
    df["Code Produit en Cuve"] = df["Code Produit en Cuve"].apply(normalize_str)
    for c in ["Stock initial (L)", "Consommé cumul (L)", "Stock restant (L)"]:
        df[c] = coerce_float(df[c]).fillna(0)
    return df

def read_existing_journal(ledger_xlsx):
    try:
        xls = pd.ExcelFile(ledger_xlsx)
        if "JOURNAL" in xls.sheet_names:
            j = pd.read_excel(ledger_xlsx, sheet_name="JOURNAL")
            return j
        return None
    except Exception:
        return None

def read_consumption_from_assemblage(assemblage_xlsx, essai="E1"):
    header_row = find_header_row(assemblage_xlsx, needle="Clé Produit en Cuve")
    if header_row is None:
        raise ValueError("Impossible de trouver l'en-tête du tableau dans l'assemblage (Clé Produit en Cuve).")

    df = pd.read_excel(assemblage_xlsx, header=header_row)
    qty_col = f"Quantité utilisée {essai}"

    if "Code Produit en Cuve" not in df.columns:
        raise ValueError("Colonne 'Code Produit en Cuve' introuvable dans l'assemblage.")
    if qty_col not in df.columns:
        raise ValueError(f"Colonne '{qty_col}' introuvable dans l'assemblage.")

    d = df.copy()
    d["Code Produit en Cuve"] = d["Code Produit en Cuve"].apply(normalize_str)
    d[qty_col] = coerce_float(d[qty_col])

    d = d[(d["Code Produit en Cuve"] != "") & (d["Code Produit en Cuve"].str.upper() != "SOUS-TOTAL")]
    d = d[d[qty_col].fillna(0) > 0].copy()

    cons = (
        d.groupby("Code Produit en Cuve", as_index=False)
        .agg(**{"Consommé (L)": (qty_col, "sum")})
        .sort_values("Code Produit en Cuve")
    )
    cons["Consommé (L)"] = cons["Consommé (L)"].round(2)
    return cons

def apply_consumption(ledger_df, cons_df):
    led = ledger_df.copy()
    cons = cons_df.copy()

    merged = led.merge(cons, on="Code Produit en Cuve", how="outer")
    merged["Stock initial (L)"] = merged["Stock initial (L)"].fillna(0)
    merged["Consommé cumul (L)"] = merged["Consommé cumul (L)"].fillna(0)
    merged["Stock restant (L)"] = merged["Stock restant (L)"].fillna(merged["Stock initial (L)"] - merged["Consommé cumul (L)"])
    merged["Consommé (L)"] = merged["Consommé (L)"].fillna(0)

    merged["Stock restant après (L)"] = merged["Stock restant (L)"] - merged["Consommé (L)"]
    merged["Surconsommation (L)"] = np.where(merged["Stock restant après (L)"] < -1e-9, -merged["Stock restant après (L)"], 0)

    merged["Consommé cumul (L)"] = merged["Consommé cumul (L)"] + merged["Consommé (L)"]
    merged["Stock restant (L)"] = merged["Stock restant après (L)"]

    updated = merged.drop(columns=["Stock restant après (L)"]).copy()
    updated = updated[["Code Produit en Cuve", "Stock initial (L)", "Consommé cumul (L)", "Stock restant (L)", "Consommé (L)", "Surconsommation (L)"]]
    updated = updated.sort_values("Code Produit en Cuve")
    return updated

def export_stock_with_highlight_and_journal(updated_df, cons_df, journal_df, ref_assemblage, date_conso):
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        out_path = tmp.name

    stock_sheet = "STOCK_MAJ"
    recap_sheet = "RECAP_CONSO"
    journal_sheet = "JOURNAL"

    recap = cons_df.copy()
    total = float(recap["Consommé (L)"].sum()) if not recap.empty else 0.0
    recap = pd.concat(
        [
            recap,
            pd.DataFrame([{"Code Produit en Cuve": "TOTAL", "Consommé (L)": round(total, 2)}]),
        ],
        ignore_index=True
    )

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        updated_df.drop(columns=["Surconsommation (L)"], errors="ignore").to_excel(writer, sheet_name=stock_sheet, index=False)
        recap.to_excel(writer, sheet_name=recap_sheet, index=False)
        if journal_df is not None:
            journal_df.to_excel(writer, sheet_name=journal_sheet, index=False)

    wb = load_workbook(out_path)
    ws = wb[stock_sheet]
    ws2 = wb[recap_sheet]
    ws3 = wb[journal_sheet] if journal_sheet in wb.sheetnames else None

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    changed_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.insert_rows(1)
    ws.merge_cells("A1:E1")
    ws["A1"] = f"Stock mis à jour — {ref_assemblage} — {date_conso.strftime('%d/%m/%Y')}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    for cell in ws[2]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 14

    header_map = {ws.cell(row=2, column=c).value: c for c in range(1, ws.max_column + 1)}
    cons_col = header_map.get("Consommé (L)")

    for r in range(3, ws.max_row + 1):
        cons_val = ws.cell(row=r, column=cons_col).value if cons_col else 0
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = border
            if c != 1:
                cell.number_format = "0.00"
        if cons_col and cons_val and float(cons_val) > 0:
            for c in range(1, ws.max_column + 1):
                ws.cell(row=r, column=c).fill = changed_fill

    ws2.insert_rows(1)
    ws2.merge_cells("A1:B1")
    ws2["A1"] = "RÉCAP des quantités utilisées (assemblage)"
    ws2["A1"].font = Font(bold=True, size=13)
    ws2["A1"].alignment = Alignment(horizontal="center")

    for cell in ws2[2]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    ws2.column_dimensions["A"].width = 24
    ws2.column_dimensions["B"].width = 16

    for r in range(3, ws2.max_row + 1):
        ws2[f"A{r}"].border = border
        ws2[f"B{r}"].border = border
        ws2[f"B{r}"].number_format = "0.00"
        if ws2[f"A{r}"].value == "TOTAL":
            ws2[f"A{r}"].font = Font(bold=True)
            ws2[f"B{r}"].font = Font(bold=True)
            ws2[f"A{r}"].fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
            ws2[f"B{r}"].fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")

    if ws3 is not None:
        for cell in ws3[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = border
        for r in range(2, ws3.max_row + 1):
            for c in range(1, ws3.max_column + 1):
                ws3.cell(row=r, column=c).border = border

    wb.save(out_path)
    return out_path

# ==========================================================
# ONGLET 1 : ASSEMBLAGE (COMPLET)
# ==========================================================
def tab_assemblage():
    with st.sidebar:
        st.header("🧪 Assemblage — imports")
        uploaded_file_cuves = st.file_uploader("État cuverie (Excel)", type=["xlsx"], key="cuves_ass")
        uploaded_file_codes = st.file_uploader("Codes produits (Excel)", type=["xlsx"], key="codes_ass")
        uploaded_file_codes_assemblage = st.file_uploader("Liste produits ASSEMBLAGE (Excel)", type=["xlsx"], key="ass_list_ass")

        st.divider()
        st.header("🧪 Assemblage — paramètres")
        st.number_input("Nombre d'essais", min_value=1, max_value=10, value=5, step=1, key="essais_ass")
        st.text_input("Titre du fichier", value="Assemblage Avril 2025", key="titre_ass")

    if not (uploaded_file_cuves and uploaded_file_codes and uploaded_file_codes_assemblage):
        st.info("👉 Importer les 3 fichiers (cuverie + codes + liste assemblage) dans la sidebar.")
        return

    df_cuves = pd.read_excel(uploaded_file_cuves)
    df_codes = pd.read_excel(uploaded_file_codes)
    df_codes_ass = pd.read_excel(uploaded_file_codes_assemblage)

    # Normalisation
    if "Produit" in df_cuves.columns:
        df_cuves["Produit"] = norm_str_series(df_cuves["Produit"])
    if "Cépage" in df_cuves.columns:
        df_cuves["Cépage"] = norm_str_series(df_cuves["Cépage"])

    # Robustesse N° Cuve
    if "N° Cuve" in df_cuves.columns:
        df_cuves["N° Cuve"] = df_cuves["N° Cuve"].apply(normalize_cuve_number)
        df_cuves["_cuve_int"] = df_cuves["N° Cuve"].apply(cuve_to_int_or_none)
        df_cuves = df_cuves[df_cuves["_cuve_int"].notna()].copy()
        df_cuves["N° Cuve"] = df_cuves["_cuve_int"].astype(int)
        df_cuves.drop(columns=["_cuve_int"], inplace=True)

    for d in (df_codes, df_codes_ass):
        for col in ["Code Produit en Cuve", "Clé Produit en Cuve", "Libéllé Produit en Cuve"]:
            if col in d.columns:
                d[col] = norm_str_series(d[col])

    # Check colonnes
    for col in ["En Stock", "Année", "Produit", "N° Cuve", "Cépage"]:
        if col not in df_cuves.columns:
            st.error(f"Le fichier cuverie doit contenir la colonne '{col}'.")
            return

    # Filtre stock
    df_cuves = df_cuves[df_cuves["En Stock"] > 0].copy()

    # Assemblages
    if "Code Produit en Cuve" not in df_codes_ass.columns:
        st.error("Le fichier ASSEMBLAGE doit contenir une colonne 'Code Produit en Cuve'.")
        return
    set_assemblages = set(df_codes_ass["Code Produit en Cuve"].dropna().astype(str).str.strip().tolist())

    # Type année/réserve
    df_cuves["Type"] = df_cuves["Année"].apply(lambda x: "Vin de l'année" if x >= 2025 else "Vin de réserve")

    df_cuves_ass = df_cuves[df_cuves["Produit"].isin(set_assemblages)].copy()
    df_cuves_std = df_cuves[~df_cuves["Produit"].isin(set_assemblages)].copy()
    df_cuves_std["CépageCode"] = df_cuves_std["Cépage"].apply(to_cepage_code)

    # UI sélection
    c1, c2, c3 = st.columns(3)
    c1.metric("Cuves en stock", int(len(df_cuves)))
    c2.metric("Cuves standard", int(df_cuves_std["N° Cuve"].nunique()))
    c3.metric("Cuves assemblage", int(df_cuves_ass["N° Cuve"].nunique()))

    st.subheader("🧩 Sélection des cuves")
    st.caption("Sélectionne des cuves par catégorie. Tu peux cumuler Assemblage + cépages classiques.")

    cuves_selectionnees = []

    if not df_cuves_ass.empty:
        with st.expander("🧩 ASSEMBLAGE", expanded=True):
            cuves_ass = st.multiselect(
                "Sélectionner les cuves ASSEMBLAGE",
                options=df_cuves_ass["N° Cuve"].tolist(),
                format_func=lambda x: (
                    f"{x} - {df_cuves_ass.loc[df_cuves_ass['N° Cuve'] == x, 'Produit'].values[0]} "
                    f"({df_cuves_ass.loc[df_cuves_ass['N° Cuve'] == x, 'En Stock'].values[0]} L)"
                ),
                key="assemblages_select"
            )
            cuves_selectionnees.extend(cuves_ass)

    cepage_codes = [c for c in df_cuves_std["CépageCode"].dropna().unique().tolist() if str(c).strip() != ""]
    if df_cuves_std.empty:
        st.warning("Aucune cuve standard (hors assemblage) trouvée en stock.")
        return

    def sort_key(x):
        order = {"C": 1, "N": 2, "M": 3}
        return (order.get(str(x).upper(), 99), str(x))

    for code in sorted(cepage_codes, key=sort_key):
        label = CEPAGE_LABEL.get(str(code).upper(), str(code))
        with st.expander(f"🍇 {label}", expanded=True):
            df_cepage = df_cuves_std[df_cuves_std["CépageCode"] == code].copy()

            df_annee = df_cepage[df_cepage["Type"] == "Vin de l'année"]
            if not df_annee.empty:
                st.markdown("**🟢 Vin de l'année (>= 2025)**")
                cuves_annee = st.multiselect(
                    f"{label} - Vin de l'année",
                    options=df_annee["N° Cuve"].tolist(),
                    format_func=lambda x, d=df_annee: (
                        f"{x} - {d.loc[d['N° Cuve'] == x, 'Produit'].values[0]} "
                        f"({d.loc[d['N° Cuve'] == x, 'En Stock'].values[0]} L)"
                    ),
                    key=f"{code}_annee_select"
                )
                cuves_selectionnees.extend(cuves_annee)

            df_reserve = df_cepage[df_cepage["Type"] == "Vin de réserve"]
            if not df_reserve.empty:
                st.markdown("**🟡 Vins de réserve (< 2025)**")
                cuves_reserve = st.multiselect(
                    f"{label} - Réserve",
                    options=df_reserve["N° Cuve"].tolist(),
                    format_func=lambda x, d=df_reserve: (
                        f"{x} - {d.loc[d['N° Cuve'] == x, 'Produit'].values[0]} "
                        f"({d.loc[d['N° Cuve'] == x, 'En Stock'].values[0]} L - {d.loc[d['N° Cuve'] == x, 'Année'].values[0]})"
                    ),
                    key=f"{code}_reserve_select"
                )
                cuves_selectionnees.extend(cuves_reserve)

    st.divider()
    st.write(f"✅ **Cuves sélectionnées : {len(set(cuves_selectionnees))}**")

    if not cuves_selectionnees:
        st.info("👉 Sélectionne au moins une cuve (standard ou assemblage) pour générer le fichier.")
        return

    df_selection = df_cuves[df_cuves["N° Cuve"].isin(cuves_selectionnees)].copy()

    # Fusion codes
    df_codes_all = pd.concat([df_codes, df_codes_ass], ignore_index=True)
    needed_cols = {"Code Produit en Cuve", "Clé Produit en Cuve", "Libéllé Produit en Cuve"}
    missing = needed_cols - set(df_codes_all.columns)
    if missing:
        st.error(
            f"Le fichier codes doit contenir : {', '.join(sorted(needed_cols))}. "
            f"Manquantes : {', '.join(sorted(missing))}"
        )
        return

    df_codes_all = df_codes_all.drop_duplicates(subset=["Code Produit en Cuve"], keep="first")

    df_selection = df_selection.merge(
        df_codes_all[["Code Produit en Cuve", "Clé Produit en Cuve", "Libéllé Produit en Cuve"]],
        how="left",
        left_on="Produit",
        right_on="Code Produit en Cuve"
    )

    # Préparer la liste de cuves pour dégustation
    cuves_for_tasting = (
        df_selection[["N° Cuve", "Produit"]]
        .drop_duplicates()
        .sort_values(["N° Cuve", "Produit"])
        .apply(lambda r: f"{r['N° Cuve']} - {str(r['Produit']).strip()}", axis=1)
        .tolist()
    )
    st.session_state["last_cuves_for_tasting"] = cuves_for_tasting
    st.session_state["last_titre_excel"] = st.session_state.get("titre_ass", "Assemblage")

    # Catégorie couleur (C/N/M/ASSEMBLAGE)
    df_selection["Catégorie couleur"] = df_selection.apply(
        lambda r: "ASSEMBLAGE" if str(r["Produit"]).strip() in set_assemblages else to_cepage_code(r["Cépage"]),
        axis=1
    )

    # Cépage affiché (texte)
    df_selection["Cépage_aff"] = df_selection["Catégorie couleur"].apply(
        lambda c: "Assemblage" if str(c).strip().upper() == "ASSEMBLAGE" else CEPAGE_LABEL.get(str(c).strip().upper(), str(c))
    )

    def _annee_export(row):
        if str(row["Produit"]).strip() in set_assemblages:
            return ""
        try:
            return int(row["Année"])
        except Exception:
            return row["Année"]

    df_selection["Année_export"] = df_selection.apply(_annee_export, axis=1)

    df_selection["Is_reserve"] = df_selection.apply(
        lambda r: 0 if str(r["Produit"]).strip() in set_assemblages else (1 if r["Type"] == "Vin de réserve" else 0),
        axis=1
    )

    df_export = df_selection[[
        "Clé Produit en Cuve",
        "N° Cuve",
        "Produit",
        "Libéllé Produit en Cuve",
        "Cépage_aff",
        "Année_export",
        "En Stock",
        "Catégorie couleur",
        "Is_reserve",
    ]].copy()

    df_export.columns = [
        "Clé Produit en Cuve",
        "N° Cuve",
        "Code Produit en Cuve",
        "Libellé Produit en Cuve",
        "Cépage",
        "Année",
        "Volume_base",
        "Catégorie couleur",
        "Is_reserve",
    ]

    df_export["__cat_order"] = df_export["Catégorie couleur"].apply(lambda x: 1 if str(x).strip().upper() != "ASSEMBLAGE" else 3)
    df_export["__reserve_order"] = df_export["Is_reserve"].apply(lambda x: 2 if int(x) == 1 else 1)

    def annee_sort(v):
        v = str(v).strip()
        if v == "":
            return -999999
        try:
            return int(v)
        except Exception:
            return -999999

    df_export["__annee"] = df_export["Année"].apply(annee_sort)

    df_export = (
        df_export.sort_values(
            by=["__cat_order", "__reserve_order", "Cépage", "__annee", "Code Produit en Cuve", "N° Cuve"],
            ascending=[True, True, True, False, True, True]
        )
        .drop(columns=["__cat_order", "__reserve_order", "__annee", "Is_reserve"])
    )

    df_sommaire = (
        df_export.groupby(["Cépage", "Année"], dropna=False)
        .agg(Nb_Cuves=("N° Cuve", "nunique"), Volume_L=("Volume_base", "sum"))
        .reset_index()
    )

    def libelle_bloc(cepage, annee):
        c = str(cepage).strip()
        a = str(annee).strip()
        return "ASSEMBLAGE" if c.lower() == "assemblage" else f"{c} {a}"

    df_sommaire["Libellé"] = df_sommaire.apply(lambda r: libelle_bloc(r["Cépage"], r["Année"]), axis=1)
    df_sommaire = df_sommaire[["Libellé", "Nb_Cuves", "Volume_L"]]
    df_sommaire = pd.concat([df_sommaire, pd.DataFrame([{
        "Libellé": "TOTAL",
        "Nb_Cuves": int(df_sommaire["Nb_Cuves"].sum()),
        "Volume_L": float(df_sommaire["Volume_L"].sum())
    }])], ignore_index=True)

    base_cols = ["Clé Produit en Cuve", "N° Cuve", "Code Produit en Cuve", "Libellé Produit en Cuve", "Cépage", "Année"]
    essais_cols = []
    ESSAIS = int(st.session_state.get("essais_ass", 5))
    for e in range(1, ESSAIS + 1):
        c = essai_cols(e)
        essais_cols.extend([c["vol"], c["solde"], c["qty"], c["pct"], c["c250"], c["c500"]])

    df_final = pd.DataFrame(columns=base_cols + essais_cols + ["Catégorie couleur"])

    for (cepage_aff, annee, cat_color), group in df_export.groupby(["Cépage", "Année", "Catégorie couleur"], sort=False):
        titre = "ASSEMBLAGE" if str(cepage_aff).strip().lower() == "assemblage" else f"{cepage_aff} {annee}"
        titre_row = {col: "" for col in df_final.columns}
        titre_row["Clé Produit en Cuve"] = titre
        titre_row["Catégorie couleur"] = cat_color
        df_final = pd.concat([df_final, pd.DataFrame([titre_row])], ignore_index=True)

        rows = []
        for _, r in group.iterrows():
            row = {col: "" for col in df_final.columns}
            row["Clé Produit en Cuve"] = r["Clé Produit en Cuve"]
            row["N° Cuve"] = r["N° Cuve"]
            row["Code Produit en Cuve"] = r["Code Produit en Cuve"]
            row["Libellé Produit en Cuve"] = r["Libellé Produit en Cuve"]
            row["Cépage"] = r["Cépage"]
            row["Année"] = r["Année"]
            row["Catégorie couleur"] = r["Catégorie couleur"]
            for e in range(1, ESSAIS + 1):
                cc = essai_cols(e)
                row[cc["vol"]] = r["Volume_base"]
            rows.append(row)

        df_final = pd.concat([df_final, pd.DataFrame(rows)], ignore_index=True)

        st_row = {col: "" for col in df_final.columns}
        st_row["Clé Produit en Cuve"] = "Sous-total"
        st_row["Année"] = annee
        st_row["Catégorie couleur"] = cat_color
        df_final = pd.concat([df_final, pd.DataFrame([st_row])], ignore_index=True)

    # =========================
    # EXCEL OUTPUT
    # =========================
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        fichier_excel = tmp.name
        df_final.to_excel(fichier_excel, index=False)

        wb = load_workbook(fichier_excel)
        ws = wb.active

        start_base_col = 7
        last_visible_col = 6 + ESSAIS * 6
        last_visible_letter = get_column_letter(last_visible_col)

        tech_col_idx = last_visible_col + 1
        tech_col_letter = get_column_letter(tech_col_idx)

        cuv_col = "B"
        ann_col = "F"
        titre_excel = st.session_state.get("titre_ass", "Assemblage")

        ws.insert_rows(1)
        ws.merge_cells(f"A1:{last_visible_letter}1")
        ws["A1"] = titre_excel
        ws["A1"].font = Font(bold=True, size=16)
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 26

        ws.column_dimensions[tech_col_letter].hidden = True

        fill_vert = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        fill_rouge = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        fill_gris = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        fill_sous_total = PatternFill(start_color="595959", end_color="595959", fill_type="solid")
        fill_header = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        font_header = Font(bold=True, color="FFFFFF")

        border_epaisse = Border(
            top=Side(border_style="thick", color="000000"),
            left=Side(border_style="thick", color="000000"),
            right=Side(border_style="thick", color="000000"),
            bottom=Side(border_style="thick", color="000000")
        )

        n = len(df_sommaire)
        ws.insert_rows(2, amount=(1 + 1 + n + 1))
        ws["A2"] = "SOMMAIRE"
        ws["A2"].font = Font(bold=True, size=12)

        ws["A3"], ws["B3"], ws["C3"] = "Catégorie", "Nb cuves", "Volume (L)"
        for cell in (ws["A3"], ws["B3"], ws["C3"]):
            cell.fill = fill_header
            cell.font = font_header

        start = 4
        for ridx, row in enumerate(df_sommaire.itertuples(index=False), start=start):
            ws[f"A{ridx}"] = row.Libellé
            ws[f"B{ridx}"] = row.Nb_Cuves
            ws[f"C{ridx}"] = row.Volume_L
            ws[f"C{ridx}"].number_format = "0.00"

        after_summary_row = start + n
        header_row = after_summary_row + 1
        data_start_row = header_row + 1
        ws.freeze_panes = f"A{data_start_row}"

        for c in range(1, last_visible_col + 1):
            cell = ws[f"{get_column_letter(c)}{header_row}"]
            cell.fill = fill_header
            cell.font = font_header
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[header_row].height = 20

        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 10
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 34
        ws.column_dimensions["E"].width = 14
        ws.column_dimensions["F"].width = 8

        for e in range(1, ESSAIS + 1):
            cols = excel_cols_for_essai(e, start_base_col=start_base_col)
            ws.column_dimensions[get_column_letter(cols["vol"])].width = 12
            ws.column_dimensions[get_column_letter(cols["solde"])].width = 11
            ws.column_dimensions[get_column_letter(cols["qty"])].width = 18
            ws.column_dimensions[get_column_letter(cols["pct"])].width = 9
            ws.column_dimensions[get_column_letter(cols["c250"])].width = 9
            ws.column_dimensions[get_column_letter(cols["c500"])].width = 9

        for r in range(data_start_row, ws.max_row + 1):
            ws[f"B{r}"].number_format = "0"

        force_excel_years_to_int(ws, year_col_letter="F", cuve_col_letter="B", start_row=data_start_row, end_row=ws.max_row)

        current_start = data_start_row
        for r in range(data_start_row, ws.max_row + 1):
            if ws[f"A{r}"].value == "Sous-total":
                for e in range(1, ESSAIS + 1):
                    cols = excel_cols_for_essai(e, start_base_col=start_base_col)
                    for col_idx in range(cols["vol"], cols["end"] + 1):
                        col_letter = get_column_letter(col_idx)
                        ws[f"{col_letter}{r}"].value = f"=SUM({col_letter}{current_start}:{col_letter}{r-1})"
                        if col_idx == cols["pct"]:
                            ws[f"{col_letter}{r}"].number_format = "0.00%"
                        elif col_idx in (cols["c250"], cols["c500"]):
                            ws[f"{col_letter}{r}"].number_format = "0"
                        else:
                            ws[f"{col_letter}{r}"].number_format = "0.00"
                current_start = r + 1

        recap_rows = [
            ("RÉCAP % - 2025 (Vin de l'année)", None),
            ("Chardonnay", "C"),
            ("Pinot Noir", "N"),
            ("Meunier", "M"),
            ("Assemblage", "ASSEMBLAGE"),
            ("TOTAL 2025", "__TOTAL_2025__"),
            ("RÉCAP % - Total (avec réserves)", None),
            ("Chardonnay", "C"),
            ("Pinot Noir", "N"),
            ("Meunier", "M"),
            ("Assemblage", "ASSEMBLAGE"),
            ("% RÉSERVE", "__RESERVE__"),
            ("TOTAL GLOBAL", "__TOTAL_ALL__"),
        ]
        recap_start_row = ws.max_row + 1

        for k, (label, key) in enumerate(recap_rows):
            rr = recap_start_row + k
            ws[f"A{rr}"] = label
            ws[f"A{rr}"].font = Font(bold=True) if (key is None or str(key).startswith("__")) else Font(bold=False)

            for e in range(1, ESSAIS + 1):
                cols = excel_cols_for_essai(e, start_base_col=start_base_col)
                qty_col_letter = get_column_letter(cols["qty"])
                pct_col_letter = get_column_letter(cols["pct"])

                cuve_criteria = f'{cuv_col}:{cuv_col},">0"'
                denom_2025 = f"SUMIFS({qty_col_letter}:{qty_col_letter},{cuve_criteria},{ann_col}:{ann_col},2025)"
                denom_all = f"SUMIFS({qty_col_letter}:{qty_col_letter},{cuve_criteria})"

                if key is None:
                    ws[f"{pct_col_letter}{rr}"] = ""
                    continue

                if key == "__TOTAL_2025__":
                    ws[f"{pct_col_letter}{rr}"] = f"=IF({denom_2025}=0,0,1)"
                    ws[f"{pct_col_letter}{rr}"].number_format = "0.00%"
                    continue

                if key == "__TOTAL_ALL__":
                    ws[f"{pct_col_letter}{rr}"] = f"=IF({denom_all}=0,0,1)"
                    ws[f"{pct_col_letter}{rr}"].number_format = "0.00%"
                    continue

                if key == "__RESERVE__":
                    num_reserve = (
                        f"SUMIFS({qty_col_letter}:{qty_col_letter},"
                        f'{cuv_col}:{cuv_col},">0",'
                        f'{ann_col}:{ann_col},">0",'
                        f'{ann_col}:{ann_col},"<2025")'
                    )
                    ws[f"{pct_col_letter}{rr}"] = f"=IF({denom_all}=0,0,{num_reserve}/{denom_all})"
                    ws[f"{pct_col_letter}{rr}"].number_format = "0.00%"
                    continue

                is_block_2025 = rr < recap_start_row + 6
                if is_block_2025:
                    num = (
                        f"SUMIFS({qty_col_letter}:{qty_col_letter},"
                        f'{cuv_col}:{cuv_col},">0",'
                        f'{tech_col_letter}:{tech_col_letter},"{key}",'
                        f'{ann_col}:{ann_col},2025)'
                    )
                    ws[f"{pct_col_letter}{rr}"] = f"=IF({denom_2025}=0,0,{num}/{denom_2025})"
                else:
                    num = (
                        f"SUMIFS({qty_col_letter}:{qty_col_letter},"
                        f'{cuv_col}:{cuv_col},">0",'
                        f'{tech_col_letter}:{tech_col_letter},"{key}")'
                    )
                    ws[f"{pct_col_letter}{rr}"] = f"=IF({denom_all}=0,0,{num}/{denom_all})"
                ws[f"{pct_col_letter}{rr}"].number_format = "0.00%"

        dernier_row = ws.max_row + 1
        ws[f"A{dernier_row}"] = "TOTAL"
        ws[f"A{dernier_row}"].font = Font(bold=True, size=12)
        ws[f"A{dernier_row}"].fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
        ws.merge_cells(start_row=dernier_row, start_column=1, end_row=dernier_row, end_column=6)

        for e in range(1, ESSAIS + 1):
            cols = excel_cols_for_essai(e, start_base_col=start_base_col)
            for col_idx in range(cols["vol"], cols["end"] + 1):
                col_letter = get_column_letter(col_idx)
                refs = []
                for rr in range(data_start_row, dernier_row):
                    if ws[f"A{rr}"].value == "Sous-total":
                        refs.append(f"{col_letter}{rr}")
                if refs:
                    ws[f"{col_letter}{dernier_row}"] = f"=SUM({','.join(refs)})"
                    ws[f"{col_letter}{dernier_row}"].font = Font(bold=True)
                    if col_idx == cols["pct"]:
                        ws[f"{col_letter}{dernier_row}"].number_format = "0.00%"
                    elif col_idx in (cols["c250"], cols["c500"]):
                        ws[f"{col_letter}{dernier_row}"].number_format = "0"
                    else:
                        ws[f"{col_letter}{dernier_row}"].number_format = "0.00"

        # Couleurs C/N/M + sous-total
        for r in range(data_start_row, ws.max_row + 1):
            cat = str(ws[f"{tech_col_letter}{r}"].value).strip().upper()
            if cat == "C":
                for cell in ws[r]:
                    cell.fill = fill_vert
            elif cat == "N":
                for cell in ws[r]:
                    cell.fill = fill_rouge
            elif cat == "M" or "ASSEMBLAGE" in cat:
                for cell in ws[r]:
                    cell.fill = fill_gris

            if ws[f"A{r}"].value == "Sous-total":
                for cell in ws[r]:
                    cell.fill = fill_sous_total
                    cell.font = Font(bold=True, color="FFA500")

        for r in range(data_start_row, ws.max_row + 1):
            val = ws[f"A{r}"].value
            cat_val = ws[f"{tech_col_letter}{r}"].value
            if val and val not in ("Sous-total", "TOTAL") and ws[f"B{r}"].value in (None, "") and cat_val not in (None, ""):
                ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=last_visible_col)
                ws[f"A{r}"].font = Font(bold=True, size=13)
                ws[f"A{r}"].alignment = Alignment(horizontal="left", vertical="center")

        for e in range(1, ESSAIS + 1):
            cols = excel_cols_for_essai(e, start_base_col=start_base_col)
            left_letter = get_column_letter(cols["start"])
            right_letter = get_column_letter(cols["end"])
            for r in range(header_row, ws.max_row + 1):
                cellL = ws[f"{left_letter}{r}"]
                cellL.border = Border(left=Side(border_style="thick", color="000000"),
                                      right=cellL.border.right, top=cellL.border.top, bottom=cellL.border.bottom)
                cellR = ws[f"{right_letter}{r}"]
                cellR.border = Border(right=Side(border_style="thick", color="000000"),
                                      left=cellR.border.left, top=cellR.border.top, bottom=cellR.border.bottom)

        groupe_debut = None
        for r in range(data_start_row, ws.max_row + 1):
            if ws[f"A{r}"].value not in ("Sous-total", "TOTAL", None, "") and ws[f"B{r}"].value not in (None, ""):
                if groupe_debut is None:
                    groupe_debut = r
            if ws[f"A{r}"].value == "Sous-total" and groupe_debut:
                for row in ws.iter_rows(min_row=groupe_debut, max_row=r, min_col=1, max_col=7):
                    for cell in row:
                        cell.border = border_epaisse
                groupe_debut = None

        total_row = ws.max_row
        for r in range(data_start_row, total_row):
            if ws[f"A{r}"].value == "Sous-total":
                continue
            if ws[f"B{r}"].value in (None, ""):
                continue

            for e in range(1, ESSAIS + 1):
                cols = excel_cols_for_essai(e, start_base_col=start_base_col)
                vol = get_column_letter(cols["vol"])
                solde = get_column_letter(cols["solde"])
                qty = get_column_letter(cols["qty"])
                pct = get_column_letter(cols["pct"])
                c250 = get_column_letter(cols["c250"])
                c500 = get_column_letter(cols["c500"])

                ws[f"{solde}{r}"].value = f"={vol}{r}-{qty}{r}"
                ws[f"{pct}{r}"].value = f"=IF({qty}{total_row}=0,0,{qty}{r}/{qty}{total_row})"
                ws[f"{pct}{r}"].number_format = "0.00%"

                ws[f"{c250}{r}"].value = f"={pct}{r}*250"
                ws[f"{c500}{r}"].value = f"={pct}{r}*500"
                ws[f"{c250}{r}"].number_format = "0"
                ws[f"{c500}{r}"].number_format = "0"

        wb.save(fichier_excel)

    st.success("✅ Fichier assemblage prêt !")
    with open(fichier_excel, "rb") as f:
        st.download_button("📥 Télécharger l'assemblage (Excel)", f, file_name="assemblage.xlsx", use_container_width=True)

    # Création essai dégustation + partage lien + PIN
    st.divider()
    st.subheader("🍷 Dégustation (optionnel) — création + partage")

    if not supabase_ready():
        st.info("Supabase non configuré (ajoute SUPABASE_URL + SUPABASE_SERVICE_ROLE_KEY dans st.secrets).")
        return

    cuves_for_tasting = st.session_state.get("last_cuves_for_tasting", [])
    if not cuves_for_tasting:
        st.info("Aucune cuve en session pour créer un essai.")
        return

    titre_excel = st.session_state.get("titre_ass", "Assemblage")
    default_essai_name = f"{titre_excel} — {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    essai_name = st.text_input("Nom de l'essai dégustation", value=default_essai_name, key="essai_name_auto")

    base_url_default = st.secrets.get("APP_URL", "")  # ex: https://tonapp.streamlit.app
    base_url = st.text_input("URL de l'app (pour générer le lien à partager)", value=base_url_default, help="Optionnel mais recommandé. Exemple: https://xxx.streamlit.app")

    if st.button("✅ Créer l'essai dégustation (Supabase) + Générer PIN", type="primary"):
        essai_id, pin = sup_create_essai(essai_name, cuves_for_tasting)
        st.session_state["deg_essai_id"] = essai_id
        st.session_state["deg_pin_last"] = pin
        sup_fetch_notes.clear()

        st.success("Essai créé ✅")
        st.markdown(f"**PIN (à transmettre aux dégustateurs) :** <span class='mono'><b>{pin}</b></span>", unsafe_allow_html=True)

        # Lien à partager (essai pré-sélectionné via query params)
        if base_url.strip():
            q = urlencode({"essai": essai_id})
            share_link = base_url.strip().rstrip("/") + "/?" + q
            st.markdown("**Lien dégustation à partager :**")
            st.code(share_link)
            st.caption("Les dégustateurs ouvrent le lien, puis saisissent le PIN pour accéder à l’essai.")
        else:
            st.caption("👉 Renseigne l’URL de l’app (APP_URL dans Secrets) pour générer automatiquement le lien.")

# ==========================================================
# ONGLET 2 : DEGUSTATION LIVE (SUPABASE) + COMPARAISON
# (CORRIGÉ : radio au lieu de tabs + keys uniques plotly)
# ==========================================================
def tab_degustation_live():
    st.subheader("🍷 Dégustation Live (multi-dégustateurs, consolidation, PIN)")

    if not supabase_ready():
        st.error("Supabase non configuré. Ajoute SUPABASE_URL et SUPABASE_SERVICE_ROLE_KEY dans les Secrets Streamlit.")
        st.stop()

    st.markdown(
        """
        <div class="card">
          <div><b>Principe :</b> chaque note est enregistrée en base (Supabase). Le dashboard se met à jour automatiquement.</div>
          <div class="small"><b>Pureté</b> = 6 - Défaut (pour que “plus grand = mieux”).</div>
        </div>
        """,
        unsafe_allow_html=True
    )

    # Query param: essai
    qp = st.query_params
    qp_essai = qp.get("essai", "")

    # Charger liste essais
    df_ess = sup_list_essais()
    if df_ess.empty:
        st.info("Aucun essai. Crée-en un depuis l’onglet Assemblage.")
        st.stop()

    df_ess["label"] = df_ess.apply(lambda r: f"{r['nom']}  —  {r['created_at']}", axis=1)

    # Choix essai (pré-sélection via query param, sinon session, sinon 0)
    current = st.session_state.get("deg_essai_id", "")
    ids = df_ess["id"].tolist()

    default_idx = 0
    if qp_essai and qp_essai in ids:
        default_idx = ids.index(qp_essai)
    elif current and current in ids:
        default_idx = ids.index(current)

    chosen_label = st.selectbox("Essai à utiliser", df_ess["label"].tolist(), index=default_idx, key="essai_select")
    chosen_id = df_ess.loc[df_ess["label"] == chosen_label, "id"].values[0]
    st.session_state["deg_essai_id"] = chosen_id

    essai = sup_get_essai(chosen_id)
    cuves = essai.get("cuves", []) or []
    st.caption(f"Essai : **{essai.get('nom','')}** — {len(cuves)} cuve(s)")

    # PIN gate (par essai)
    if "authorized_essais" not in st.session_state:
        st.session_state["authorized_essais"] = set()

    if chosen_id not in st.session_state["authorized_essais"]:
        with st.expander("🔐 Accès dégustateurs (PIN)", expanded=True):
            pin_default = st.session_state.get("deg_pin_last", "") if qp_essai else ""
            pin = st.text_input(
                "PIN",
                value=pin_default,
                type="password",
                help="Demande le PIN au créateur de l’essai.",
                key=f"pin_{chosen_id}",
            )
            if st.button("Déverrouiller l’essai", key=f"unlock_{chosen_id}"):
                if sup_check_pin(essai, pin):
                    st.session_state["authorized_essais"].add(chosen_id)
                    st.success("OK ✅ Essai déverrouillé")
                    st.rerun()
                else:
                    st.error("PIN incorrect.")
        st.stop()

    # ✅ IMPORTANT : ne plus utiliser st.tabs ici (les 2 blocs s'exécutent à chaque rerun).
    view = st.radio(
        "Vue",
        ["📝 Saisie dégustateur", "📡 Dashboard live + comparaison"],
        horizontal=True,
        key=f"deg_view_{chosen_id}",
    )

    # ---------------------------
    # VUE 1: Saisie (FORM + retry + verrou)
    # ---------------------------
    if view == "📝 Saisie dégustateur":
        if "saving_note" not in st.session_state:
            st.session_state["saving_note"] = False
        if "current_cuve_idx" not in st.session_state:
            st.session_state["current_cuve_idx"] = 0

        colA, colB, colC = st.columns([1.1, 1.1, 0.8])
        with colA:
            degustateur = st.text_input(
                "Dégustateur",
                value=st.session_state.get("degustateur", ""),
                key=f"degustateur_{chosen_id}",
            )
            if degustateur:
                st.session_state["degustateur"] = degustateur

        with colB:
            if cuves:
                idx = int(st.session_state.get("current_cuve_idx", 0)) % max(1, len(cuves))
                cuve = st.selectbox("Cuve", cuves, index=idx, key=f"cuve_select_{chosen_id}")
            else:
                cuve = st.selectbox("Cuve", ["(aucune)"], key=f"cuve_select_{chosen_id}_none")

        with colC:
            if cuves:
                st.caption("Navigation cuves")
                if st.button("➡️ Cuve suivante", use_container_width=True, key=f"next_cuve_{chosen_id}"):
                    st.session_state["current_cuve_idx"] = (int(st.session_state["current_cuve_idx"]) + 1) % len(cuves)
                    st.rerun()

        with st.form(f"form_note_{chosen_id}", clear_on_submit=False):
            c1, c2, c3 = st.columns(3)
            with c1:
                acidite = st.slider("Acidité", 1, 5, 3, key=f"acidite_{chosen_id}")
                amertume = st.slider("Amertume", 1, 5, 3, key=f"amertume_{chosen_id}")
            with c2:
                mineralite = st.slider("Minéralité", 1, 5, 3, key=f"mineralite_{chosen_id}")
                volume = st.slider("Volume", 1, 5, 3, key=f"volume_{chosen_id}")
            with c3:
                sucrosite = st.slider("Sucrosité", 1, 5, 3, key=f"sucrosite_{chosen_id}")
                defaut = st.slider("Défaut (1=aucun, 5=fort)", 1, 5, 1, key=f"defaut_{chosen_id}")

            commentaire = st.text_area("Commentaire", height=90, key=f"comment_{chosen_id}")

            submitted = st.form_submit_button(
                "✅ Envoyer / Mettre à jour",
                disabled=st.session_state["saving_note"],
                use_container_width=True
            )

        if submitted:
            if not degustateur or not degustateur.strip():
                st.error("Renseigne le nom du dégustateur.")
            elif not cuves or cuve in ("(aucune)", "", None):
                st.error("Aucune cuve disponible.")
            else:
                st.session_state["saving_note"] = True
                with st.spinner("Enregistrement..."):
                    ok, err = sup_upsert_note_with_retry(
                        essai_id=chosen_id,
                        cuve=cuve,
                        degustateur=degustateur.strip(),
                        notes={
                            "acidite": int(acidite),
                            "amertume": int(amertume),
                            "mineralite": int(mineralite),
                            "volume": int(volume),
                            "sucrosite": int(sucrosite),
                            "defaut": int(defaut),
                        },
                        commentaire=commentaire,
                        tries=3
                    )

                st.session_state["saving_note"] = False

                if ok:
                    sup_fetch_notes.clear()
                    st.success("Note enregistrée ✅")
                    st.rerun()
                else:
                    st.error(f"Erreur Supabase (réseau/timeout possible) : {err}")

    # ---------------------------
    # VUE 2: Dashboard (autorefresh uniquement ici)
    # ---------------------------
    else:
        left, right = st.columns([1, 1])
        with left:
            refresh_s = st.slider("Refresh (secondes)", 1, 10, 2, key=f"refresh_{chosen_id}")
        with right:
            by_taster = st.checkbox(
                "Araignées : afficher aussi par dégustateur",
                value=False,
                key=f"by_taster_{chosen_id}",
            )

        # ✅ autorefresh uniquement quand on est sur le dashboard
        st_autorefresh(interval=int(refresh_s) * 1000, key=f"deg_live_refresh_{chosen_id}")

        df = sup_fetch_notes(chosen_id)
        if df.empty:
            st.info("Aucune note pour le moment.")
            st.stop()

        st.caption(f"{len(df)} note(s) — {df['cuve'].nunique()} cuve(s) — {df['degustateur'].nunique()} dégustateur(s)")

        cuves_with_data = [c for c in cuves if c in set(df["cuve"].dropna())]
        cuves_no_data = [c for c in cuves if c not in set(df["cuve"].dropna())]
        if cuves_no_data:
            st.caption("Sans notes : " + ", ".join(cuves_no_data))

        st.markdown("### 🕸️ Araignées par cuve")
        for i in range(0, len(cuves_with_data), 3):
            cols = st.columns(3)
            for j in range(3):
                if i + j >= len(cuves_with_data):
                    break
                cuv = cuves_with_data[i + j]
                dcuve = df[df["cuve"] == cuv]
                with cols[j]:
                    st.markdown(f"**Cuve : {cuv}**  \nVotes : {len(dcuve)}")
                    fig = radar_fig(dcuve, by_taster=by_taster)
                    st.plotly_chart(
                        fig,
                        use_container_width=True,
                        key=f"radar_{chosen_id}_{cuv}_{'t' if by_taster else 'm'}",
                    )

        st.markdown("### 📊 Comparaison cuves (moyennes consolidées)")
        scores = build_scores_table(df)
        st.dataframe(scores, use_container_width=True)

        hm = plot_heatmap_scores(scores)
        if hm is not None:
            st.plotly_chart(hm, use_container_width=True, key=f"hm_{chosen_id}")

        st.markdown("### 📏 Mesurer l’écart entre cuves")
        st.caption("Distance = écart global sur les 6 critères (Acidité, Amertume, Minéralité, Volume, Sucrosité, Pureté).")
        ref_cuve = st.selectbox("Choisir une cuve de référence", scores["cuve"].astype(str).tolist(), key=f"ref_{chosen_id}")
        dist_fig = plot_distance_to_reference(scores, ref_cuve)
        if dist_fig is not None:
            st.plotly_chart(dist_fig, use_container_width=True, key=f"dist_{chosen_id}_{ref_cuve}")

        with st.expander("💬 Commentaires"):
            st.dataframe(
                df[["created_at", "cuve", "degustateur", "commentaire"]].sort_values("created_at", ascending=False),
                use_container_width=True
            )

# ==========================================================
# ONGLET 3 : STOCK UPDATE (ANTI DOUBLE + DELTA + EXPORT)
# ==========================================================
def tab_stock_update():
    st.subheader("📦 Mise à jour des stocks (anti double-application + rapprochement)")
    st.markdown(
        """
        <div class="card">
          <div><b>But :</b> importer l'état de stock + l'assemblage validé (essai choisi) et décrémenter le stock par <b>Code Produit en Cuve</b>.</div>
          <div class="small">Anti double-application : même fichier + essai + ref + date = bloqué si déjà appliqué (via Fingerprint).</div>
          <div class="small">Rapprochement : compare l'état réel (snapshot) vs le ledger théorique.</div>
        </div>
        """,
        unsafe_allow_html=True
    )

    col1, col2, col3 = st.columns([1.2, 1.2, 1])
    with col1:
        stock_file = st.file_uploader("📥 État des stocks (Excel)", type=["xlsx"], key="stock_file")
        st.caption("Colonnes attendues : Produit, En Stock")
    with col2:
        assemblage_file = st.file_uploader("🧪 Assemblage validé (Excel)", type=["xlsx"], key="ass_valid")
        st.caption("On lit 'Code Produit en Cuve' + 'Quantité utilisée E1/E2...'")
    with col3:
        ledger_file = st.file_uploader("🗂️ Optionnel : Ledger précédent (export de l'app)", type=["xlsx"], key="ledger_prev")
        st.caption("Permet de cumuler plusieurs assemblages.")

    st.divider()
    a, b, c, d = st.columns([1, 1, 1, 1.2])
    with a:
        essai = st.selectbox("Essai à appliquer", ["E1", "E2", "E3", "E4", "E5"], index=0)
    with b:
        date_conso = st.date_input("Date", value=datetime.now().date())
    with c:
        ref_assemblage = st.text_input("Référence", value=f"Assemblage_{datetime.now().strftime('%Y%m%d')}")
    with d:
        st.caption("💡 Mets la référence du lot validé (ou n° de fiche).")

    if not stock_file or not assemblage_file:
        st.info("👉 Importer état des stocks + assemblage validé.")
        return

    try:
        df_stock = pd.read_excel(stock_file)
        snapshot = build_stock_snapshot(df_stock)

        existing_journal = None
        if ledger_file:
            ledger = read_ledger(ledger_file)
            existing_journal = read_existing_journal(ledger_file)
            st.success("Ledger existant chargé ✅")
        else:
            ledger = init_ledger_from_snapshot(snapshot)
            existing_journal = None
            st.warning("Aucun ledger fourni → initialisation depuis l'état des stocks.")

        assemblage_bytes = assemblage_file.getvalue()
        fp = make_fingerprint(assemblage_bytes, essai=essai, ref=ref_assemblage, date_conso=date_conso)

        if journal_has_fingerprint(existing_journal, fp):
            st.error("⛔ Cet assemblage a déjà été appliqué au stock (même fichier + essai + ref + date).")
            st.stop()

        cons = read_consumption_from_assemblage(assemblage_file, essai=essai)

        x1, x2, x3 = st.columns(3)
        x1.metric("Codes consommés", int(len(cons)))
        x2.metric("Total consommé (L)", float(cons["Consommé (L)"].sum()) if not cons.empty else 0.0)
        x3.metric("Fingerprint", fp)

        st.markdown("### 🧾 Récap consommation (par code produit)")
        st.dataframe(cons, use_container_width=True)

        updated = apply_consumption(ledger, cons)
        issues = updated[updated["Surconsommation (L)"] > 0].copy()

        if not issues.empty:
            st.error("⛔ Surconsommation détectée : dépassement de stock restant !")
            st.dataframe(issues[["Code Produit en Cuve", "Stock restant (L)", "Consommé (L)", "Surconsommation (L)"]], use_container_width=True)
            return

        st.success("✅ OK : stock mise à jour possible.")
        st.markdown("### 📌 Aperçu stock à jour")
        st.dataframe(updated.drop(columns=["Surconsommation (L)"]), use_container_width=True)

        journal_new = cons.copy()
        journal_new["Date"] = pd.to_datetime(date_conso)
        journal_new["Référence"] = ref_assemblage
        journal_new["Essai"] = essai
        journal_new["Fingerprint"] = fp
        journal_new = journal_new[["Date", "Référence", "Essai", "Fingerprint", "Code Produit en Cuve", "Consommé (L)"]]

        if existing_journal is not None and not existing_journal.empty:
            journal_all = pd.concat([existing_journal, journal_new], ignore_index=True)
        else:
            journal_all = journal_new

        st.markdown("### 🔎 Rapprochement : État des stocks vs Ledger")
        delta = build_delta_table(snapshot, updated)

        seuil = st.number_input("Seuil d'alerte écart (L)", min_value=0.0, value=50.0, step=10.0)
        delta_alert = delta[delta["Écart (Etat - Ledger)"].abs() >= float(seuil)]

        d1, d2 = st.columns(2)
        d1.metric("Nb codes (écart >= seuil)", int(len(delta_alert)))
        d2.metric("Écart max (L)", float(delta["Écart (Etat - Ledger)"].abs().max()) if not delta.empty else 0.0)

        st.dataframe(delta.head(200), use_container_width=True)

        if not delta_alert.empty:
            st.warning("⚠️ Des écarts importants existent (transferts, pertes, snapshot différent du théorique).")
            st.dataframe(delta_alert.head(200), use_container_width=True)

        out_path = export_stock_with_highlight_and_journal(
            updated_df=updated,
            cons_df=cons,
            journal_df=journal_all,
            ref_assemblage=ref_assemblage,
            date_conso=date_conso
        )

        with open(out_path, "rb") as f:
            st.download_button(
                "📥 Télécharger STOCK_MAJ (surligné) + RECAP_CONSO + JOURNAL",
                f,
                file_name=f"STOCK_MAJ_{ref_assemblage}.xlsx",
                use_container_width=True
            )

    except Exception as e:
        st.error(f"Erreur : {e}")

# ==========================================================
# ROUTING
# ==========================================================
tab1, tab2, tab3 = st.tabs(["🧪 Assemblage", "🍷 Dégustation Live", "📦 Mise à jour stocks"])

with tab1:
    tab_assemblage()

with tab2:
    tab_degustation_live()

with tab3:
    tab_stock_update()
